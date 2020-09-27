from openpyxl import load_workbook
from openpyxl import Workbook
import os
import datos

def leerUnaHoja(hoja):
    sheet = ''
    try:
        workbook = load_workbook(datos.excelBook, read_only=False)
        sheet = workbook[hoja]

        arraySheet = []
        for row in sheet.iter_rows():
            arraySheet.append(row)

        return arraySheet
    except:
        print('Nombre de hoja incorrecto. \n==== EJECUCIÓN FINALIZADA ====')
        exit()

def guardarDatos(numeros, hoja):
    i = 0
    workbook = load_workbook(datos.excelBook, read_only=False)
    sheet = workbook[hoja]
    for tel in numeros:
        i += 1
        celda = sheet.cell(row = i, column = 3)
        celda.value = tel

    workbook.save(datos.excelBook)


def obtenerNombresSheets():
    workbook = load_workbook(datos.excelBook, read_only=True)
    return workbook.sheetnames

def seleccionarLibro():
    try:
        print('SELECCIONAR LIBROOO')
        directorio = 'datos/'
        contenido = os.listdir(directorio)
        for x in range(len(contenido)):
            print(str (x) + '- ' + contenido[x])
        
        libro = int(input('\nSeleccione archivo:  '))
        print('\n\n > Libro seleccionado: ' + contenido[libro])
        datos.excelBook = datos.directorioDatos + '/' + contenido[libro]
    except:
        print('\nERROR EN LECTURA DE LIBRO')

    