from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook

def todo():
    listaTelefonos = []
    workbook = load_workbook('datos.xlsx', read_only=False)
    sheet = workbook['hoja']
    driver = webdriver.Chrome(executable_path="drivers/chromedriver")
    driver.get('https://www.telexplorer.cl/')
    for row in sheet.iter_rows():
        print(row[0].value)
        actionChains = ActionChains(driver)

        direccion = driver.find_element_by_xpath(".//a[text()='Dirección']")
        direccion.click()
        direccion.click()


        #wait = WebDriverWait(driver, 15)
        #element = wait.until(EC.presence_of_element_located((By.XPATH, ".//a[text()='Dirección']")))

        actionChains.double_click(direccion).perform()


        buscar = driver.find_element_by_id('boton_buscar_3')
        calle = driver.find_element_by_id('calle')
        altura = driver.find_element_by_id('altura')
        provincia = Select(driver.find_element_by_id('provincia2'))

        waitd = WebDriverWait(driver, 15)
        elementd = waitd.until(EC.visibility_of(calle))

        calle.send_keys(row[0].value)
        altura.send_keys(row[1].value)
        provincia.select_by_value('10')
        buscar.click()

        try:
            localidad = Select(driver.find_element_by_id('localidad'))
            localidad.select_by_value('1060419')

            listaTelefonos.append(driver.find_element_by_class_name('resultado_telefono').text)
        except:
            listaTelefonos.append('-')

        driver.get('https://www.telexplorer.cl/')

    return listaTelefonos

def guardarDatos(numeros):
    i = 0
    workbook = load_workbook('datos.xlsx', read_only=False)
    sheet = workbook['hoja']
    for tel in numeros:
        i += 1
        celda = sheet.cell(row = i, column = 3)
        celda.value = tel

    workbook.save('datos.xlsx') 

listado = todo()
guardarDatos(listado)
print(listado)



